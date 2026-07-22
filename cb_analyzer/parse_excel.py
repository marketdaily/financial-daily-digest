"""把『CB發行案件更新』Excel 解析成結構化 cb_database.json。
換新檔只要 python parse_excel.py <新檔.xlsx>，系統其他部分不用改。"""
import sys
import os
import re
import json
import datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(HERE, "cb_database.json")
DEFAULT_XLSX = "/mnt/c/Users/USER/Downloads/CB發行案件更新_20260626.xlsx"


def _s(c):
    if c is None:
        return ""
    if isinstance(c, datetime.datetime):
        return c.strftime("%Y-%m-%d")
    return str(c).strip()


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _parse_premium(raw):
    """溢價率%：可能是 '104.36' / '106~120' / '105.1~110' / '未定' / ''。
    回傳 (low, high, mid)；單值時三者相同；未定回 (None,None,None)。"""
    raw = (raw or "").strip().replace("～", "~").replace(" ", "")
    if not raw or raw == "未定":
        return None, None, None
    if "~" in raw:
        a, b = raw.split("~", 1)
        lo, hi = _num(a), _num(b)
        if lo is None or hi is None:
            return None, None, None
        return lo, hi, (lo + hi) / 2
    v = _num(raw)
    return (v, v, v) if v is not None else (None, None, None)


def _parse_tcri(raw):
    """'TCRI6/無擔' → (6, '無擔')；'TCRI7/土地銀' → (7, '土地銀')。"""
    raw = raw or ""
    m = re.search(r"TCRI\s*(\d+)", raw)
    rating = int(m.group(1)) if m else None
    coll = ""
    if "/" in raw:
        coll = raw.split("/", 1)[1].strip()
    secured = bool(coll) and coll not in ("無擔", "無擔保")
    return rating, coll, secured


def _parse_put(raw):
    """賣回條件 'YTP(2)=(0%)' / 'YTM(3)=(0%~0.5%)' / 'YTP(2)=(0.5%~1.5%)' / '未定'。
    回傳 dict: kind(YTP/YTM), year, yield_low, yield_high, yield_mid (年化%)。"""
    raw = (raw or "").replace("～", "~").replace(" ", "")
    if not raw or raw == "未定":
        return {"kind": None, "year": None, "y_low": None, "y_high": None, "y_mid": None, "raw": raw}
    m = re.search(r"(YTP|YTM)\((\d+)\)=\(([^)]*)\)", raw)
    if not m:
        return {"kind": None, "year": None, "y_low": None, "y_high": None, "y_mid": None, "raw": raw}
    kind, year, ystr = m.group(1), int(m.group(2)), m.group(3)
    nums = [_num(x.replace("%", "")) for x in re.split(r"[~\-]", ystr) if x]
    nums = [n for n in nums if n is not None]
    if not nums:
        y_low = y_high = y_mid = 0.0
    else:
        y_low, y_high = min(nums), max(nums)
        y_mid = (y_low + y_high) / 2
    return {"kind": kind, "year": year, "y_low": y_low, "y_high": y_high, "y_mid": y_mid, "raw": raw}


def _parse_year(raw):
    m = re.search(r"(\d+)", raw or "")
    return int(m.group(1)) if m else None


def _parse_split_date(raw):
    """ASO可拆解日 '2026/07/07 (二)' → '2026-07-07'。"""
    m = re.search(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", raw or "")
    if not m:
        return None
    return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"


def _parse_clearing(note):
    """從備註抓真實清算價(佔面額100):
      競拍 → 承銷價X(最準);沒有就用(低標+高標)/2。
      詢圈/固定 → '依票面金額之N%發行' 的 N。
    回傳 dict: clearing(最佳估計), auction_low, auction_high, method, fixed_pct。"""
    note = (note or "").replace("，", ",")
    low = high = sold = fixed = None
    m = re.search(r"承銷價\s*(\d{2,3}(?:\.\d+)?)", note)
    if m:
        sold = _num(m.group(1))
    m = re.search(r"低標\s*(\d{2,3}(?:\.\d+)?)", note)
    if m:
        low = _num(m.group(1))
    m = re.search(r"高標\s*(\d{2,3}(?:\.\d+)?)", note)
    if m:
        high = _num(m.group(1))
    # 詢圈/固定發行:依(票面金額|面額)之 N% 發行
    m = re.search(r"(?:票面金額|面額)[之]?\s*(\d{2,3}(?:\.\d+)?)\s*%", note)
    if m:
        v = _num(m.group(1))
        if v and 90 <= v <= 200:
            fixed = v
    if sold:
        clearing, method = sold, "競拍承銷價"
    elif low and high:
        clearing, method = (low + high) / 2, "競拍中值"
    elif fixed:
        clearing, method = fixed, "固定發行"
    else:
        clearing, method = None, "未定"
    return {"clearing": clearing, "auction_low": low, "auction_high": high,
            "fixed_pct": fixed, "method": method}


def parse(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    import parse_masterlink
    if parse_masterlink.is_masterlink(ws):          # 元富『CB初級市場資訊』格式自動委派
        return parse_masterlink.parse_ws(ws)
    rows = list(ws.iter_rows(values_only=True))
    section = ""
    out = []
    for r in rows[1:]:
        sect_cell = _s(r[0])
        if sect_cell:
            section = sect_cell.replace("\n", "")
        code = _s(r[1])
        name = _s(r[3])
        if not code or not name:
            continue
        p_lo, p_hi, p_mid = _parse_premium(_s(r[11]))
        rating, coll, secured = _parse_tcri(_s(r[4]))
        put = _parse_put(_s(r[15]))
        note = _s(r[17])
        clr = _parse_clearing(note)
        out.append({
            "section": section,
            "stock_code": code,
            "bond_code": _s(r[2]),
            "name": name,
            "tcri": rating,
            "collateral": coll,
            "secured": secured,
            "size_yi": _num(_s(r[5])),       # 發行量(億)
            "underwriter": _s(r[6]),
            "announce_date": _s(r[7]) or None,
            "effective_date": _s(r[9]) or None,
            "bookbuild": _s(r[10]),          # 詢圈/競拍
            "premium_low": p_lo,             # 轉換溢價率%（轉換價/訂價參考股價）
            "premium_high": p_hi,
            "premium_mid": p_mid,
            "conv_price": _num(_s(r[12])),   # 轉換價
            "listing_date": _s(r[13]) or None,
            "split_date": _parse_split_date(_s(r[14])),  # 可拆解日
            "put": put,
            "tenor_year": _parse_year(_s(r[16])),
            "clearing_price": clr["clearing"],     # 真實清算/承銷價(競拍或固定)
            "auction_low": clr["auction_low"],
            "auction_high": clr["auction_high"],
            "pricing_method": clr["method"],
            "issue_price": clr["clearing"],         # 相容舊欄位
            "note": note,
        })
    return out


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    data = parse(xlsx)
    db = {
        "source_file": os.path.basename(xlsx),
        "parsed_at": None,  # 由呼叫端蓋日期，避免 Date.now 之類不可重現
        "count": len(data),
        "items": data,
    }
    with open(DB_PATH, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)
    priced = sum(1 for d in data if d["conv_price"] and d["premium_mid"])
    print(f"解析 {len(data)} 檔 → {DB_PATH}")
    print(f"  已定價(有轉換價+溢價率): {priced} 檔")
    print(f"  條件未定: {len(data) - priced} 檔")


if __name__ == "__main__":
    main()
